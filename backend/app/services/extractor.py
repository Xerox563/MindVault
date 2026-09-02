import os
from typing import Optional
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from app.utils.logger import log_error, log_info

def extract_text(file_path: str, file_type: str) -> Optional[str]:
    ext = file_type.lower()
    try:
        if ext == ".pdf":
            result = extract_pdf(file_path)
        elif ext == ".docx":
            result = extract_docx(file_path)
        elif ext == ".xlsx":
            result = extract_xlsx(file_path)
        elif ext in (".txt", ".csv", ".md"):
            result = extract_txt(file_path)
        else:
            log_error(f"Unsupported file type: {ext}")
            return None
        if result:
            log_info(f"Extracted {len(result)} chars from {file_path}")
        return result
    except Exception as e:
        log_error(f"Extraction failed for {file_path}: {str(e)}")
        return None

def extract_pdf(file_path: str) -> Optional[str]:
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text() or ""
    return text if text.strip() else None

def extract_docx(file_path: str) -> Optional[str]:
    doc = Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs if para.text])

def extract_xlsx(file_path: str) -> Optional[str]:
    wb = load_workbook(file_path)
    text = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " ".join([str(cell) for cell in row if cell])
            if row_text:
                text.append(row_text)
    return "\n".join(text) if text else None

def extract_txt(file_path: str) -> Optional[str]:
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()
